import datetime
import re
from decimal import Decimal
from sqlalchemy.orm import Session

from app.models.models import (
    Project, Employee, ProjectMember, Payment, BusinessManager
)
from app.services.bonus_calculator import calculate_bonus_pool


def _patch_openpyxl():
    from openpyxl.worksheet.datavalidation import DataValidation
    _orig = DataValidation.__init__
    def _patched(self, *args, **kwargs):
        for k in ('id', 'showDropDown', 'showErrorMessage', 'showInputMessage'):
            kwargs.pop(k, None)
        _orig(self, *args, **kwargs)
    DataValidation.__init__ = _patched


def import_from_excel(filepath: str, db: Session) -> dict:
    _patch_openpyxl()
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    stats = {"projects": 0, "sub_projects": 0, "employees": 0, "payments": 0}

    # ═══ STEP 1: Import 人员工资 FIRST ═══
    if "人员工资" in wb.sheetnames:
        ws = wb["人员工资"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h else "" for h in rows[0]]
        col = {h: i for i, h in enumerate(headers)}

        for row in rows[1:]:
            vals = [str(v) if v is not None else "" for v in row]
            if all(v == "" for v in vals):
                continue
            name = _g(vals, col, "姓名")
            if not name or name == "合计":
                continue

            emp = db.query(Employee).filter(Employee.name == name).first()
            if not emp:
                emp = Employee(name=name)
                db.add(emp)
                db.flush()
                stats["employees"] += 1

            emp.position = _g(vals, col, "工种")
            emp.level = _g(vals, col, "级别")
            is_active = _g(vals, col, "在职状态") == "在职"
            emp.status = "在职" if is_active else "离职"
            zz = _g(vals, col, "是否转正")
            emp.employment_type = zz if zz in ("正式", "试用", "实习") else "试用"
            emp.salary = _dec(_g(vals, col, "工资")) or Decimal("0")
            emp.guarantee = _dec(_g(vals, col, "保底")) or Decimal("0")
            emp.social_security = _dec(_g(vals, col, "社保")) or Decimal("0")
            hire_date = _date(_g(vals, col, "入职时间"))
            if hire_date and not emp.hire_date:
                emp.hire_date = hire_date
            if not is_active:
                remark = _g(vals, col, "备注")
                if remark and not emp.remark:
                    emp.remark = remark
        db.commit()

    # Ensure 原粼 has an employee record (not in salary sheet but is an employee)
    yl = db.query(Employee).filter(Employee.name == "原粼").first()
    if not yl:
        db.add(Employee(name="原粼", status="在职", employment_type="正式"))
        db.flush()
        stats["employees"] += 1
    db.commit()

    # ═══ STEP 2: Import project sheets (skip 一部遗留项目) ═══
    for sheet_name in ["项目提成核算表"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h else "" for h in rows[0]]
        col = {h: i for i, h in enumerate(headers)}

        last_parent_id = None

        for row in rows[1:]:
            vals = [str(v) if v is not None else "" for v in row]
            if all(v == "" for v in vals):
                continue

            project_name = _g(vals, col, "项目名称")
            if not project_name:
                continue

            amount = _dec(_g(vals, col, "项目金额")) or Decimal("0")
            is_sub = project_name.startswith("↘")
            clean_name = project_name.lstrip("↘")

            # Check existing
            existing = db.query(Project).filter(
                Project.project_name == clean_name,
                Project.deleted_at.is_(None)
            ).first()
            if existing:
                proj = existing
            else:
                proj = Project(project_name=clean_name)
                db.add(proj)
                db.flush()
                if is_sub:
                    stats["sub_projects"] += 1
                else:
                    stats["projects"] += 1

            proj.amount = amount
            proj.region = _g(vals, col, "地区")
            proj.parent_project_id = last_parent_id if is_sub else None

            # Status auto-advance
            acceptance = _date(_g(vals, col, "验收时间"))
            delivery = _date(_g(vals, col, "实际交付日期"))
            if acceptance:
                proj.status = "已分成"
            elif delivery:
                proj.status = "已交付"
            elif _g(vals, col, "UI确认时间"):
                proj.status = "UI确认"
            else:
                proj.status = "未开始"

            proj.contract_date = _date(_g(vals, col, "项目签单时间"))
            proj.ui_confirm_date = _date(_g(vals, col, "UI确认时间"))
            proj.actual_delivery_date = delivery
            proj.project_cycle_month = _dec(_g(vals, col, "项目周期（月）")) or Decimal("1")

            # ── Business Manager ──
            bm_name = _g(vals, col, "商务经理")
            if bm_name:
                bm_names = re.split(r"[/／]", bm_name)
                for bmn in bm_names:
                    bmn = bmn.strip()
                    if not bmn:
                        continue
                    # Strip percentage suffixes like "靳林9/陈久亿1"
                    bmn = re.sub(r"\d+$", "", bmn).strip()
                    if not bmn:
                        continue
                    bm = db.query(BusinessManager).filter(
                        BusinessManager.name == bmn,
                        BusinessManager.deleted_at.is_(None)
                    ).first()
                    if not bm:
                        bm = BusinessManager(name=bmn)
                        db.add(bm)
                        db.flush()
                    proj.business_manager_id = bm.id
                    break

            # ── Developers ──
            dev_raw = _g(vals, col, "开发人员")
            uiue_raw = _g(vals, col, "UIUE")
            dev_names = _parse_names(dev_raw)
            uiue_names = _parse_names(uiue_raw)

            if not dev_names and not uiue_names:
                dev_names = ["未指定"]

            total_members = len(dev_names) + len(uiue_names) or 1
            bonus_pool = calculate_bonus_pool(amount)
            per_person_bonus = bonus_pool / total_members

            product_bonus = _dec(_g(vals, col, "产品提成金额")) or Decimal("0")
            pb_per_person = product_bonus / total_members

            proj_month = _dec(_g(vals, col, "项目周期（月）")) or Decimal("1")

            # Create project members (skip if already exist)
            existing_members = db.query(ProjectMember).filter(
                ProjectMember.project_id == proj.id,
                ProjectMember.deleted_at.is_(None)
            ).first()

            if not existing_members:
                for dev_name in dev_names:
                    emp = _ensure_employee(dev_name, db)
                    if emp is None:
                        continue  # Skip non-employees
                    db.add(ProjectMember(
                        project_id=proj.id, employee_id=emp.id,
                        role="开发", input_month=proj_month,
                        bonus=per_person_bonus, product_bonus=Decimal("0"),
                    ))
                for ui_name in uiue_names:
                    emp = _ensure_employee(ui_name, db)
                    if emp is None:
                        continue
                    db.add(ProjectMember(
                        project_id=proj.id, employee_id=emp.id,
                        role="UI/UE", input_month=proj_month,
                        bonus=per_person_bonus, product_bonus=pb_per_person,
                    ))

            # ── Payments ──
            existing_payment = db.query(Payment).filter(
                Payment.project_id == proj.id,
                Payment.deleted_at.is_(None)
            ).first()

            if not existing_payment:
                for i in range(1, 5):
                    pay_val = _g(vals, col, f"已回款{i}")
                    if pay_val and pay_val != "~":
                        pay_amount = _dec(pay_val)
                        if pay_amount and pay_amount > 0:
                            db.add(Payment(
                                project_id=proj.id,
                                payment_amount=pay_amount,
                                payment_stage=i,
                                status="已到账",
                                actual_payment_date=datetime.date.today(),
                            ))
                            stats["payments"] += 1

            if not is_sub:
                last_parent_id = proj.id

        db.commit()

    wb.close()
    return stats


def _parse_names(raw: str) -> list[str]:
    """Parse developer names: '张云稳+李耀男' → ['张云稳','李耀男']; '原粼→李耀男' → ['李耀男']"""
    if not raw:
        return []
    # Handle "→" (hand-off): take the last person
    if "→" in raw:
        parts = raw.split("→")
        raw = parts[-1].strip()
    # Split by "+" or "＋"
    names = []
    for part in re.split(r"[+＋]", raw):
        name = part.strip()
        # Filter out percentage/ratio suffixes
        name = re.sub(r"\d+(\.\d+)?$", "", name).strip()
        if name:
            names.append(name)
    return names


def _ensure_employee(name: str, db: Session) -> Employee | None:
    emp = db.query(Employee).filter(Employee.name == name).first()
    if not emp:
        return None  # Not in approved employee list, skip
    return emp


def _dec(val: str) -> Decimal | None:
    if not val or val in ("~", "#REF!", ""):
        return None
    try:
        return Decimal(str(val).replace(",", "").replace("¥", ""))
    except Exception:
        return None


def _date(val: str) -> datetime.date | None:
    if not val:
        return None
    try:
        if "00:00:00" in val:
            val = val.split(" ")[0]
        dt = datetime.datetime.strptime(val, "%Y-%m-%d")
        if dt.year < 2000:
            return None
        return dt.date()
    except Exception:
        return None


def _g(vals: list, col: dict, key: str) -> str:
    idx = col.get(key, -1)
    if idx < 0 or idx >= len(vals):
        return ""
    return str(vals[idx]).strip() if vals[idx] else ""
